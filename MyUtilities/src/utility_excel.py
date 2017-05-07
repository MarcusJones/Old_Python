'''
Created on 2012-03-03

@author: Anonymous
'''
from __future__ import division
from win32com.client import Dispatch
import logging.config
import os
import re
import xlrd
import xlwt
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from utility_inspect import whoami, whosdaddy
import unittest
from config import *


class ExcelBookWrite(object):
    def __init__(self):
        #self.excelPath = excelPath

        logging.debug("Excel file at: {}, exists={}".format(self.excelPath,self.exists))

    @property
    def exists(self):
        return os.path.exists(self.excelPath)

def excelWriteTableXLSX(fullPath,sheetName,rows):
    wb = Workbook()
    sheet = wb.create_sheet()
    sheet.title = sheetName


    for i,row in enumerate(rows):
        i = i + 1

        col = get_column_letter(i)
        for j,item in enumerate(row):
            j = j + 1
            sheet.cell("{}{}".format(col,j)).value = item

    #    :
    #        sheet.write(i, j, item)

    wb.save(fullPath)
            #sheet0, 0
    logging.debug("Wrote {} rows to {}".format(len(rows),fullPath))

            #print i,row
        #print enumerate(rows)



def excelWriteTable(fullPath,sheetName,rows):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet(sheetName)
    print sheet

    for i,row in enumerate(rows):
        for j,item in enumerate(row):
            sheet.write(i, j, item)

    wb.save(fullPath)
            #sheet0, 0
    logging.debug("Wrote {} rows to {}".format(len(rows),fullPath))

            #print i,row
        #print enumerate(rows)

class ExcelBookRead(object):
    """
    import xlrd
    book = xlrd.open_workbook("myfile.xls")
    print "The number of worksheets is", book.nsheets
    print "Worksheet name(s):", book.sheet_names()
    sh = book.sheet_by_index(0)
    print sh.name, sh.nrows, sh.ncols
    print "Cell D30 is", sh.cell_value(rowx=29, colx=3)
    for rx in range(sh.nrows):
        print sh.row(rx)
    # Refer to docs for more details.
    # Feedback on API is welcomed.
    """


    def __init__(self, excelPath):
        self.excelPath = excelPath
        logging.debug("Excel file at: {}".format(self.excelPath))

    def get_sheet_names(self):
        wb = xlrd.open_workbook(self.excelPath)

        sheetNames = wb.sheet_names()
        logging.debug("Found {} sheet names: {}".format(len(sheetNames),sheetNames))
        return sheetNames

    def getTable(self, targetSheet, startRow = 0, endRow=None, startCol=0, endCol=None):

        wb = xlrd.open_workbook(self.excelPath)

        if targetSheet not in self.get_sheet_names():
            return False

        sht = wb.sheet_by_name(targetSheet)

        if not endRow:
            endRow = sht.nrows
            endRowText = "END"

        if not endCol:
            endColText = "END"

        data = list()

        #logging.debug("Looking for data, rows {} to {}, columns {} to {}".format(startRow, endRowText, startCol, endColText))

        for rowx in range(startRow,endRow):
            thisRow = sht.row_values(rowx, startCol, endCol)
            data.append(thisRow)

        #print(data)
        if len(data):
            logging.debug("Got data table from {}, {} rows, {} columns".format(targetSheet,len(data),len(data[0])))
        else:
            logging.debug("No data found in {}".format(targetSheet))

        return data

    def getTableAll(self, targetSheet, dataType="str"):

        wb = xlrd.open_workbook(self.excelPath)
        sheet = wb.sheet_by_name(targetSheet)

        data = list()
        for i in xrange(sheet.nrows):
            if dataType == "str":
                data.append(sheet.row_values(i)) #drop all the values in the rows into data
            else:
                raise Exception("Unsupported data type")

        logging.debug("Got data table from {}, {} rows, {} columns".format(targetSheet,len(data),len(data[0])))

        return data

class controlledExcelAPI:
    def __enter__(self,excelPath):
        self.object = ExcelBookAPI(excelPath)
        return self.object
    def __exit__(self, type, value, traceback):
        self.object.saveAndClose()



class ExcelBookAPI(object):


    def __enter__(self):
        logging.debug("***Enter***".format())
        return self
        #self.object = ExcelBookAPI(excelPath)
        #return self.object

    def __exit__(self, type, value, traceback):
        logging.debug("***Exit***".format())
        if self.autosave:
            self.saveAndClose()
        else:
            self.closeAll()
    def __init__(self, excelPath, autocreate = False, autosave = False):
        self.excelPath = os.path.abspath(excelPath)
        self.autosave = autosave
        logging.debug("Excel file at: {}, exists={}".format(self.excelPath,self.exists))
        self.xl = Dispatch('Excel.Application')
        self.xl.Visible = 0
        
        if not self.exists and autocreate:
            self.book = self.xl.Workbooks.Add()
            #print self.xl
            #print self.xl.Workbooks.Add()
            #print self.book
            print self.book
            print self.excelPath
            self.book.SaveAs(self.excelPath)
            logging.debug("Created file Excel file at: {}, exists={}".format(self.excelPath,self.exists))
        else:
            self.book = self.xl.Workbooks.Open(self.excelPath)

    def closeAll(self):
        #self.book.Close(0)
        self.xl.Quit()

    def saveAs(self):
        self.xl.ActiveWorkbook.SaveAs(self.excelPath)

    def save(self):
        self.xl.ActiveWorkbook.Save()

    def saveAndClose(self):
        self.xl.ActiveWorkbook.SaveAs(self.excelPath)
        self.xl.ActiveWorkbook.Close(SaveChanges=0) # see note 1
        logging.debug("Closed excel file at: {}".format(self.excelPath))

    def sheetExists(self,sheetName):
        for sheet in [name.Name for name in self.book.Sheets]:
            if sheet == sheetName:
                return True
        return False

    def getLastRow(self,sheetName):
        sh = self.book.Sheets[sheetName]
        return sh.getRows()

    def write(self,sheetName,rows,x=0,y=0):

        assert( type(rows[0]) == list or type(rows[0]) == tuple), "Need a 2D array {} = {}".format(type(rows[0]), rows[0])
        LIMIT_SHEET_NAME = 20
        if len(sheetName) > LIMIT_SHEET_NAME:
            sheetName = sheetName[0:LIMIT_SHEET_NAME]
        sheetName = sheetName.replace(":", " ")

        #print self.sheetExists(sheetName)
        #print [name.Name for name in self.book.Sheets]
        #print [name.Name for name in self.book.Worksheets]
        #prin
        if self.sheetExists(sheetName):
            # Use the existing
            sh = self.book.Sheets[sheetName]
            logging.debug("Sheet {} exists".format(sheetName))

        else:
            # Create a new, rename it
            sh = self.book.Worksheets.Add()
            #lastSheet = self.book.Sheets.Count
            #sh = self.book.Sheets[lastSheet-1]
            try:
                sh.Name = sheetName
            except:
                print sheetName
                print sheetName.__str__
                raise
            sh = self.book.Sheets[sheetName]
            logging.debug("Sheet {} created".format(sheetName))


        # Iterate over data
        for i,row in enumerate(rows):
            i += 1 + x
            
            for j,item in enumerate(row):
                j+=1 + y
                sh.Cells(i,j).Value = item

        #self.saveAndClose()

        """
        >>> from win32com.client import Dispatch
        >>> app = Dispatch('Excel.Application')
        >>> app.Visible = True
        >>> wrk = app.Workbooks.Add()
        >>> wrk.Sheets.Count
        3
        >>> sh = wrk.Sheets.Add()
        >>> wrk.Sheets.Count
        4
        >>> sh.Name
        u'Sheet4'
        >>> sh.Name = 'New Name 4'
        >>> wrk.Sheets[0].Name
        u'New Name 4'
        >>> wrk.Sheets[2].Name
        u'Sheet2'
        >>> wrk.Sheets[2].Name = 'Hello'
        >>> wrk.Save()
        >>> wrk.Close(0)
        >>> app.Quit()
        """

        """

        import win32com.client
        xlApp = win32com.client.Dispatch("Excel.Application")
        xlApp.Visible=1
        xlWb = xlApp.Workbooks.Open("Read.xls")
        print xlApp.Worksheets("Sheet1").Name
        xlApp.Worksheets("Sheet2").Range("A1").Value = "yellow"
        cell = xlApp.Worksheets("Sheet3")
        cell.Range("C3").Value = "money"
        cell.Range("D4").Value = 9999
        print cell.Range("C3").Value
        print cell.Range("D4").Value
        xlWb.Close(SaveChanges=1)
        xlApp.Quit()
        """


        #$print book
        #raise

        # create new file ('Workbook' in Excel-vocabulary)


#        # store default worksheet object so we can delete it later
#        defaultWorksheet = workbook.Worksheets(1)
#
#        # build new chart (on seperate page in workbook)
#        chart = workbook.Charts.Add()
#        chart.ChartType = constants.xlXYScatter
#        chart.Name = "Plot"
#
#        # create data worksheet
#        worksheet = workbook.Worksheets.Add()
#        worksheet.Name = "Plot data"
#
#        # install data
#        xColumn = addDataColumn(worksheet, 0, x)
#        yColumn = addDataColumn(worksheet, 1, y)
#
#        # create series for chart
#        series = chart.SeriesCollection().NewSeries()
#        series.XValues = xColumn
#        series.Values = yColumn
#        series.Name = "Data"
#        series.MarkerSize = 3
#
#
#
#

        self.save()
        logging.debug("Wrote {} rows to Excel file at: {}, sheet {}, starting at row {}".format(len(rows),self.excelPath,sheetName,x))

    @property
    def exists(self):
        return os.path.exists(self.excelPath)


    def get_sheet_names(self):
        xl = Dispatch('Excel.Application')


        sheetObjects = self.book.Worksheets
#
#        print
#
#        for sht in sheets:
#            print sht.Name
        sheets = [sht.Name for sht in sheetObjects]
        #print sheets

        logging.debug("Found {} sheet names".format(len(sheets)))

        return sheets

    def scanDown2(self, targetSheet, rowNumber, colNumber, searchString, limitScan=1000):
        """
        Pass in searchString="None" to find the next empty cell
        (As an actual string, not python None type
        Limit is 1000 rows as default
        """

        logging.debug("Scanning {}, column {}, starting row {}, for '{}'".format(targetSheet, rowNumber, colNumber, searchString))

        # Select the sheet
        sheet = self.book.Sheets(targetSheet)

        #xl.Visible = False

        # Scan down
        thisRow = rowNumber
        foundRow = None
        for row in range(rowNumber,limitScan):
            currentValue = sheet.Cells(row,colNumber).Value
            # Look for it
            #print currentValue,searchString
            #print currentValue
            #print searchString
            #print re.search(currentValue,searchString)
            currentValue = str(currentValue)
            if currentValue and searchString and re.search(searchString,currentValue):
                foundRow = thisRow
                break
            if not currentValue and not searchString:
                foundRow = thisRow
                break
            thisRow += 1

        #book.Close(SaveChanges=0) #to avoid prompt

        if foundRow:
            logging.debug("Found {} at row {}, column {}".format(searchString, foundRow, colNumber))
            return foundRow
        else:
            raise Exception("{} not found".format(searchString))

    def getTable(self,sheetName):
        table = ExcelBookRead(self.excelPath).getTable(sheetName)

        logging.debug("Got table len {} from sheet {}".format(len(table),sheetName))

        return table


    def scanDown(self, targetSheet, rowNumber, colNumber, searchString, limitScan=1000):
        """
        Pass in searchString=None to find the next empty cell
        Limit is 1000 rows as default
        """

        logging.debug("Scanning {}, column {}, starting row {}, for '{}'".format(targetSheet, rowNumber, colNumber, searchString))

        # Attach the excel COM object

        xl = Dispatch('Excel.Application')

        # Open the project file
        book = xl.Workbooks.Open(self.excelPath)

        # Select the sheet
        sheet = book.Sheets(targetSheet)

        xl.Visible = False

        # Scan down
        thisRow = rowNumber
        foundRow = None
        for row in range(rowNumber,limitScan):
            currentValue = sheet.Cells(row,colNumber).Value
            # Look for it
            #print currentValue,searchString
            #print currentValue
            #print searchString
            #print re.search(currentValue,searchString)
            currentValue = str(currentValue)
            if currentValue and searchString and re.search(searchString,currentValue):
                foundRow = thisRow
                break
            if not currentValue and not searchString:
                foundRow = thisRow
                break
            thisRow += 1

        book.Close(SaveChanges=0) #to avoid prompt

        if foundRow:
            logging.debug("Found {} at row {}, column {}".format(searchString, foundRow, colNumber))
            return foundRow
        else:
            raise Exception("{} not found".format(searchString))

    def getRows(self, targetSheet, startRow=1, endRow = 1000, startCol=1,endCol = 100):
        """
        Return cols until first blank
        """
        #logging.debug("Loading project from {0}".format(self.excelPath))

        # Attach the excel COM object

        xl = Dispatch('Excel.Application')

        # Open the project file
        book = xl.Workbooks.Open(self.excelPath)

        # Select the sheet
        sheet = book.Sheets(targetSheet)

        xl.Visible = False

        rows = list()

        if not endRow:
            runUntilRow = 1000
        else:
            runUntilRow = endRow

        if not endCol:
            runUntilCol = 100
        else:
            runUntilCol = endCol

        checks = 0

        for row in range(startRow,runUntilRow+1):
            col = 1

            #?? What is this?
            if not endRow and not sheet.Cells(row,col).Value:
                break

            # Only return non-empty rows!
            if sheet.Cells(row,col).Value is None:
                pass
            else:
                rows.append(list())

            for col in range(startCol, runUntilCol+1):
                checks += 1
                thisVal = sheet.Cells(row,col).Value
                #print checks, thisVal
                if thisVal is not None:
                    rows[-1].append(thisVal)

        book.Close(SaveChanges=0) #to avoid prompt

        logging.debug("Checked {} cells".format(checks))
        logging.debug("Returning {} rows".format(len(rows)))

        return rows

    def getCell(self, targetSheet, row, col):
        xl = Dispatch('Excel.Application')

        # Open the project file
        book = xl.Workbooks.Open(self.excelPath)

        # Select the sheet
        sheet = book.Sheets(targetSheet)

        xl.Visible = False

        thisVal = sheet.Cells(row,col).Value

        book.Close(SaveChanges=0) #to avoid prompt

        logging.debug("Returning '{}' at row {}, col {} in sheet {} ".format(thisVal, row, col,targetSheet ))

        return thisVal


    def getTable2(self, targetSheet, startRow, endRow, startCol, endCol):
        """
        startRow, Starts at 1, not 0!
        endRow, Inclusive
        startCol, Starts at 1
        endCol Inclusive
        """
        logging.debug("Loading table on {}".format(targetSheet))

        # Attach the excel COM object

        xl = Dispatch('Excel.Application')

        # Open the project file
        book = xl.Workbooks.Open(self.excelPath)

        # Select the sheet
        sheet = book.Sheets(targetSheet)

        xl.Visible = False

        rows = list()
        for row in range(startRow,endRow+1):
            rows.append(list())
            for col in range(startCol, endCol+1):
                thisVal = sheet.Cells(row,col).Value
                rows[-1].append(thisVal)

        book.Close(SaveChanges=0) #to avoid prompt

        return rows

def _test1():
    logging.debug("Started _test1".format())


    thisExcelPath = os.path.normpath(r"..\Project Definitions\testing1.xlsx")

    thisExcelPath = os.path.join(os.getcwd(), thisExcelPath)

    excelBook1 = ExcelBookAPI(thisExcelPath)

    excelBook1.scanDown("Variables", 1, 1, "Continous variables")

    excelBook1.scanDown("Variables", 7, 1, None)

    excelBook1.getTable("Variables", 1, 10, 1, 10)

    theseRows = excelBook1.getRows("Variables", 1, None, 1, None)

    print theseRows

    logging.debug("Finished _test1".format())

#===============================================================================
# Unit testing
#===============================================================================
@unittest.skip("Skip")
class allTests(unittest.TestCase):
    @unittest.skip("Skip")
    def test010(self):
        print "**** TEST {} ****".format(whoami())
        workBookPath = os.getcwd() + "\\..\\TestingFiles\\testing01.xlsx"
        wb = xlrd.open_workbook(workBookPath)
        print wb
        sh = wb.sheet_by_name(u'Sheet1')
        print sh
        for rownum in range(sh.nrows):
            print sh.row_values(rownum)

    @unittest.skip("Skip")
    def test020(self):
        print "**** TEST {} ****".format(whoami())

        workBookPath = os.getcwd() + "\\..\\TestingFiles\\testing01.xlsx"
        thisBook = ExcelBookRead(workBookPath)
        print thisBook.getTable("Sheet1", 0, 2, 0, 2)
        thisTable = thisBook.getTable("Sheet1", 0, 2)
        for row in thisTable:
            for val in row:
                print val

    def test030(self):
        print "**** TEST {} ****".format(whoami())

        testPath = r"C:\Projects\IDFout\00Test.xlsx"
        testData= [['the', 'big', 'cat', 'flies'],[3,4,5]]
        xl = ExcelBookAPI(testPath)
        xl.write("testSht",testData)

        xl.write("testSht2",testData)

        nextRow = len(xl.getTable("testSht")) + 1
        xl.write("testSht",testData, nextRow)
        xl.saveAndClose()


class test_pw(unittest.TestCase):
    def test010(self):
        print "**** TEST {} ****".format(whoami())
        workBookPath = r"C:\Users\PC1\Desktop\Test.xlsx"
        import sys
        #workBookPath = r"C:\Users\PC1\Desktop\short.xlsx"

        import win32com.client
        import itertools
        import string
        import traceback
        import datetime as dt
        import time
        xlApp = win32com.client.Dispatch("Excel.Application")
        print "Excel library version:", xlApp.Version

        for pw_size in range(5, 10):

            start_time = time.time()
            gen = itertools.combinations_with_replacement("123456789"+string.ascii_lowercase,pw_size)
            flg_found = False
            count = 0
            for password in gen:
                password = "".join(password)

                if count % 100 == 0:
                    print "{:<10} - {}".format(count, password)


                try:
                    xlwb = xlApp.Workbooks.Open(workBookPath, 0, True, None, password)
                    flg_found = True
                except:
                    pass
                    #traceback.print_exc()
                    #print "Err"
                if flg_found:
                    break

                count += 1
            end_time = time.time()
            print "Combinations length {} over {} seconds".format(pw_size,end_time - start_time)
            print "Last password was {}".format(password)
            if flg_found:
                break

        print password
        print password
        print password


if __name__ == "__main__":
    print ABSOLUTE_LOGGING_PATH
    logging.config.fileConfig(ABSOLUTE_LOGGING_PATH)


    myLogger = logging.getLogger()
    myLogger.setLevel("DEBUG")

    logging.debug("Started _main".format())

    #print FREELANCE_DIR

    unittest.main()

    logging.debug("Finished _main".format())

